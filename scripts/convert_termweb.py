import PythonTmx
from datetime import datetime
import os
import openpyxl
import openpyxl.utils
from pathlib import Path
import logging
import lxml.etree as etree

logger = logging.getLogger(__name__)

# Update language codes to match MT Glossary Software Main.xlsx exactly
LANGUAGE_CODES = {
    "English (US)": "en-US",
    "Arabic": "ar-SA",
    "Bulgarian (bg)": "bg-BG",
    "Chinese (cn)": "zh-CN",
    "Chinese (tw)": "zh-TW",
    "Croatian (hr)": "hr-HR",
    "Czech (cz)": "cs-CZ",
    "Danish (dk)": "da-DK",
    "Dutch (nl)": "nl-NL",
    "Estonian (ee)": "et-EE",
    "Finnish (fi)": "fi-FI",
    "French (fr)": "fr-FR",
    "German (de)": "de-DE",
    "Greek (gr)": "el-GR",
    "Hebrew (il)": "he-IL",
    "Hungarian (hu)": "hu-HU",
    "Indonesian (id)": "id-ID",
    "Italian (it)": "it-IT",
    "Japanese (jp)": "ja-JP",
    "Korean (kr)": "ko-KR",
    "Latvian (lv)": "lv-LV",
    "Lithuanian (lt)": "lt-LT",
    "Malay (my)": "ms-MY",
    "Norwegian Bokmål (no)": "nb-NO",
    "Polish (pl)": "pl-PL",
    "Portuguese (br)": "pt-BR",
    "Romanian (ro)": "ro-RO",
    "Russian (ru)": "ru-RU",
    "Serbian (cyrl-rs)": "sr-Cyrl-RS",
    "Slovak (sk)": "sk-SK",
    "Slovenian (si)": "sl-SI",
    "Spanish (001)": "es-ES",
    "Swedish (se)": "sv-SE",
    "Thai (th)": "th-TH",
    "Turkish (tr)": "tr-TR",
    "Ukrainian (ua)": "uk-UA",
    "Vietnamese (vn)": "vi-VN"
}

class OperationLog:
    def __init__(self):
        self.messages = []
    
    def info(self, msg):
        self.messages.append(("info", msg))
    
    def error(self, msg):
        self.messages.append(("error", msg))
    
    def get_log(self):
        return self.messages

def process_excel_file(file_path: str) -> tuple[list[str], list[tuple[str, str]]]:
    """
    Process TermWeb Excel file and convert to TMX.
    
    Args:
        file_path: Path to Excel file
    
    Returns:
        tuple: (List of created TMX files, List of (level, message) log entries)
    """
    logger.info(f"Processing Excel file: {file_path}")
    logs = []
    created_files = []
    try:
        # Load workbook with data_only=True to get values instead of formulas
        wb = openpyxl.load_workbook(file_path, data_only=True)
        

        # Initializing output path
        input_path = Path(file_path) 
        output_path = None



        # Process each worksheet
        for ws in wb.worksheets:
            
            term_dict = {}  
            logger.info(f"Processing worksheet: {ws.title}")
            
            # Create TMX for this worksheet
            segtype: PythonTmx.SEGTYPE = PythonTmx.SEGTYPE
            segtype.name = 'PHRASE'
            segtype.value = 'phrase'
            header = PythonTmx.Header(srclang="en-us",segtype=segtype,adminlang="en-us", creationtool="VATV Converter", creationtoolversion="1.0", tmf="tmx", datatype="unknown", encoding="utf8")
            tmx = PythonTmx.Tmx(header=header)
            last_column = openpyxl.utils.get_column_letter(ws.max_column) + str(1)                   #Last column of the first row, in order to obtain all target languages
            last_entry = openpyxl.utils.get_column_letter(ws.max_column) + str(ws.max_row)    #Last column of the last row, in order to obtain every cell that has content
            languages = ws['A1': last_column]        #full list of language columns, this must be accessed using [0][x], where x is the column
            terms = ws['A2': last_entry]             #full list of terms, both source and targets


            
            for row in range(len(terms[0])):                   #Loops through every row in the xlsx
                source_list = []                            #Initialize a list for all posible source values
                target_dict = {}                            #Initialize a dict for storing each target language as key and each target found as values
                for col in range(len(languages[0])):        #Loops through every cell in this ro
                    
                    if terms[row][col].value != None:
                        if languages[0][col].value == "English (US)":         #If the value is in an English column it will be stored as Source,
                            source_list.append(terms[row][col].value)                           
                        else:                                                                                          #else it will be considered as target
                            if target_dict.get(LANGUAGE_CODES.get(languages[0][col].value)):                           #it will check if a key for that target language exists, and add it as value
                                
                                target_dict[LANGUAGE_CODES.get(languages[0][col].value)].append(terms[row][col].value)
                                
                            else:                                                                                    #or it will create a new one
                                target_dict[LANGUAGE_CODES.get(languages[0][col].value)] = [terms[row][col].value,]
                              
                for source in source_list:
                    for language in target_dict:                                #then, for each source found, it will create a new tuple for each target found.
                        for translation in target_dict[language]:               #the tuple will include the source, target and target language
                            tuple_entry = (source, translation, language)
                            
                            if term_dict.get(language):                         #Generates a tuple with the source, target and target language
                                term_dict[language].append(tuple_entry)         #Checks if the dictionary contains an entry with that target language. If
                            else:                                               #it exists, adds the new entry to the list of values for that key. Otherwise, 
                                term_dict[language] = [tuple_entry,]            #creates a new entry and a new list with that element
            
            
            for key in term_dict:
                for tw_entry in term_dict[key]:                                                       #Loops through every element on the list to create a TU for each element
                    source_tuv = PythonTmx.Tuv(lang="en-us", content=tw_entry[0])                                #Assigns first element as source
                    target_tuv = PythonTmx.Tuv(lang=tw_entry[2], content=tw_entry[1])                            #Assigns second element as target
                    tmx.tus.append(PythonTmx.Tu(srclang="en-us", tuvs=[source_tuv,target_tuv]))
                output_path = input_path.parent / f"{key}.tmx"
                output_file = output_path / f"{key}.tmx"
                
                
                new_tmx_root: etree._Element = PythonTmx.to_element(tmx, True)
                etree.ElementTree(new_tmx_root).write(output_path, encoding="utf-8", xml_declaration=True)
                created_files.append(str(output_path))
                logs.append(("info", f"Created TMX for {ws.title} with {len(tmx.tus)} TUs"))
                tmx.tus = []
            # Get target language from filename
            

            # Create output path
            














        
            # Find header row and language columns
            #header_row = None
            #lang_cols = {}
            #date_cols = {}
            #language = None
            #for row in ws.iter_rows(min_row=1, max_row=10):  # Check first 10 rows for headers
            #    for cell in row:
            #        if cell.value in LANGUAGE_CODES:
            #            header_row = cell.row
            #            lang_cols[cell.column] = LANGUAGE_CODES[cell.value]
            #            language = LANGUAGE_CODES[cell.value]
            #        elif cell.value == "Creation Date":
            #            date_cols[cell.column] = "creation"
            #        elif cell.value == "Last Modified":
            #            date_cols[cell.column] = "change"
            #
            #for lang in lang_cols:
            #    print(lang_cols[lang])
            #if not header_row:
            #    logs.append(("warning", f"No language columns found in worksheet: {ws.title}"))
            #    continue
#
#
            ## Get target language from filename
            #input_path = Path(file_path)
#
            ## Create output path
            #output_path = input_path.parent / f"{language}.tmx"
#
#
#
            ## Process content rows
            #row_count = 0
            #skipped_count = 0
            #
            #for row in ws.iter_rows(min_row=header_row + 1):
            #    row_count += 1
            #    source_text = None
            #    target_texts = {}
            #    dates = {}
            #    
            #    # Get text and dates for each language
            #    for col in lang_cols:
            #        text = row[col - 1].value
            #        if text:
            #            text = str(text).strip()
            #            if lang_cols[col] == "en-US":
            #                source_text = text
            #            else:
            #                target_texts[lang_cols[col]] = text
            #    
            #    # Get dates if available
            #    for col in date_cols:
            #        date_cell = row[openpyxl.utils.column_index_from_string(col) - 1]
            #        if date_cell.value:
            #            try:
            #                if isinstance(date_cell.value, datetime):
            #                    date_str = date_cell.value.strftime("%Y%m%dT%H%M%SZ")
            #                else:
            #                    date_str = datetime.strptime(str(date_cell.value), "%Y-%m-%d").strftime("%Y%m%dT%H%M%SZ")
            #                dates[date_cols[col]] = date_str
            #            except ValueError:
            #                logger.warning(f"Invalid date format in row {row_count}: {date_cell.value}")
            #    
            #    # Create TU if we have source and at least one target
            #    if source_text and target_texts:
#
            #        # Add source TUV
            #        source_tuv = PythonTmx.Tuv(lang="en-US")
            #        source_tuv.content = source_text
            #        if "creation" in dates:
            #            source_tuv.creationdate = dates["creation"]
            #        if "change" in dates:
            #            source_tuv.changedate = dates["change"]
            #        
            #        # Add target TUVs
            #        for lang, text in target_texts.items():
            #            target_tuv = PythonTmx.Tuv(lang=lang)
            #            target_tuv.content = text
            #            if "creation" in dates:
            #                target_tuv.creationdate = dates["creation"]
            #            if "change" in dates:
            #                target_tuv.changedate = dates["change"]
#
            #        
            #        tmx.tus.append(PythonTmx.Tu(srclang="en-us", tuvs=[source_tuv,target_tuv]))
            #    else:
            #        skipped_count += 1
            
            #if tmx.tus:
            #    # Save TMX file
            #    output_file = output_path / f"{language}.tmx"
            #    new_tmx_root: etree._Element = PythonTmx.to_element(tmx, True)
            #    etree.ElementTree(new_tmx_root).write(output_path, encoding="utf-8", xml_declaration=True)
            #    created_files.append(str(output_file))
            #    logs.append(("info", f"Created TMX for {ws.title} with {len(tmx.tus)} TUs"))
            #    print("saved")
            
            
        
        return tuple(created_files)

    except Exception as e:
        logger.error(f"Error processing Excel file: {e}", exc_info=True)
        raise

def process_directory(directory):
    """
    Process all TermWeb Excel files in a directory.
    
    Args:
        directory (str): Directory containing TermWeb Excel files
    
    Returns:
        tuple: (results, log_messages)
            results: List of created TMX file paths
            log_messages: List of (level, message) tuples
    """
    logger = OperationLog()
    results = []
    
    try:
        for filename in os.listdir(directory):
            if filename.endswith(('.xlsx', '.xls')):
                source_file = os.path.join(directory, filename)
                logger.info(f"Processing file: {filename}")
                result = process_excel_file(source_file, Path(directory))
                results.extend(result[0])
        return results, logger.get_log()
    except Exception as e:
        logger.error(f"Error processing directory: {str(e)}")
        raise

if __name__ == "__main__":
    # Example usage when run directly
    directory = input("Enter directory path containing TermWeb Excel files: ")
    results, log = process_directory(directory)
    for tmx_file in results:
        print(f"Created TMX file: {tmx_file}")
    for level, message in log:
        print(f"{level.upper()}: {message}") 