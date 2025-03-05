
import PythonTmx
from datetime import datetime
import glob
import openpyxl
import openpyxl.utils

language_codes = {  "English (US)": "en-US", 
                    "Arabic": "ar-SA", 
                    "Bulgarian (bg)": "bg-BG", 
                    "Chinese (cn)": "zh-CN", 
                    "Chinese (tw)": "zh-TW", 
                    "Croatian (hr)": "hr-HR", 
                    "Czech (cz)": "cs-CZ", 
                    "Danish (dk)": "da-DK", 
                    "Dutch (nl)": "nl-NL", 
                    "Estonian (ee)": "et-EE", 
                    "Finnish (fi)": "fi-FI", 
                    "French (fr)": "fr-FR", 
                    "German (de)": "de-DE", 
                    "Greek (gr)": "el-GR", 
                    "Hebrew (il)": "he-IL", 
                    "Hungarian (hu)": "hu-HU", 
                    "Indonesian (id)": "id-ID", 
                    "Italian (it)": "it-IT", 
                    "Japanese (jp)": "ja-JP", 
                    "Korean (kr)": "ko-KR", 
                    "Latvian (lv)": "lv-LV", 
                    "Lithuanian (lt)": "lt-LT", 
                    "Malay (my)": "ms-MY", 
                    "Norwegian Bokmål (no)": "nb-NO", 
                    "Polish (pl)": "pl-PL", 
                    "Portuguese (br)": "pt-BR", 
                    "Romanian (ro)": "ro-RO", 
                    "Russian (ru)": "ru-RU", 
                    "Serbian (cyrl-rs)": "sr-Cyrl-RS", 
                    "Slovak (sk)": "sk-SK", 
                    "Slovenian (si)": "sl-SI", 
                    "Spanish (001)": "es-ES", 
                    "Swedish (se)": "sv-SE", 
                    "Thai (th)": "th-TH", 
                    "Turkish (tr)": "tr-TR", 
                    "Ukrainian (ua)": "uk-UA", 
                    "Vietnamese (vn)": "vi-VN" 
                    }

def termweb_converter(file_list):
    """
    termweb_converter(file_list: list)->dict
    
    Converts a TermWeb file into a TMX file, using the main column as source
    and the rest as targets.
    
    :param file_list: A list of TermWeb xlsx file names. 
    :return: A dict with each target language found as a key, and a list of all 
    corresponding tuples with format (source, target, target_language_code) as 
    value
    
    """
    source = None
    target = None
    language = None
    term_dict = {}                                      #A dictionary for every target lanaguage on the xlsx file

    for xlfile in file_list:                            #Loops through every xlsx file name in the list
        termweb_export = openpyxl.load_workbook(xlfile)
        term_list = termweb_export.active
        last_column = openpyxl.utils.get_column_letter(term_list.max_column) + str(1)                   #Last column of the first row, in order to obtain all target languages
        last_entry = openpyxl.utils.get_column_letter(term_list.max_column) + str(term_list.max_row)    #Last column of the last row, in order to obtain every cell that has content
        languages = term_list['A1': last_column]        #full list of language columns, this must be accessed using [0][x], where x is the column
        terms = term_list['A2': last_entry]             #full list of terms, both source and targets

        for row in range(len(terms)):                   #Loops through every row in the xlsx
            for col in range(len(languages[0])):        #Loops through every cell in this row
                if col == 0:
                    source = terms[row][col].value      #Saves the source from the first cell of the row

                elif terms[row][col].value != None:     #If it's not the first cell and it has content, it will store the value of
                    target = terms[row][col].value      #that cell and get the language code from the header of that column
                    language = languages[0][col].value
                    tuple_entry = (source, target, language_codes.get(language))    #Generates a tuple with the source, target and target language
                    if term_dict.get(language):                     #Checks if the dictionary contains an entry with that target language. If
                        term_dict[language].append(tuple_entry)     #it exists, adds the new entry to the list of values for that key. Otherwise,
                    else:                                           #creates a new entry and a new list with that element
                        term_dict[language] = [tuple_entry,]
        
        termweb_export.close()

    return term_dict                        #Returns a dictionary, where each key is a target language and 
                                            #its value is a list of (source, target)tuples for the creation of TMs        

def main():   
    termweb_exports = glob.glob("*.xlsx")                   #This list will store every file name with the ".xlsx" extension
    tuple_dict = termweb_converter(termweb_exports)         #This will convert the content of every xlsx into a dictionary, having each key as the target language
                                                            #and the value as a tuple (source, target, target_language_code)
                                                                
    for key in tuple_dict:                                  #Loops through every key in the dictionary
        tu_list = []
        for tw_entry in tuple_dict[key]:                                #Loops through every element on the list to create a TU for each element
            src_tuv = PythonTmx.Tuv(xmllang="en-us")                                #Assigns first element as source
            src_tuv.segment.text = tw_entry[0]                                      #Creates source TUV
            tgt_tuv = PythonTmx.Tuv(xmllang=tw_entry[2])                            #Assigns second element as target
            tgt_tuv.segment.text = tw_entry[1]                                      #Creates target TUV
            tu_list.append(PythonTmx.Tu(srclang="en-us", tuvs=[src_tuv,tgt_tuv]))   #Creates TU using both source and target TUVs   

        clean_tm_object = PythonTmx.Tmx(tus=tu_list)                                #Creates TMX object using the tu_list for that target language
        clean_tm_object.header = PythonTmx.Header(creationdate=datetime.today())    #Adds a header containing only the current time as creation date
        clean_tm_object.to_tmx(key + ".tmx")                                        #Creates a new TMX using the same name as the target language code
        print("\t" + key + " TM has been created")    


if __name__ == "__main__":
    main()